import pandas as pd
import re
import openpyxl
import hashlib
import csv
import os
from pathlib import Path
from typing import Union, List, Dict, Set

from ..utils.constants import (
    FileMetadata,
    STUD_ID_COL,
    DATA_CATEGORIES
)


def validate_file_naming_format(file_naming_format: list):
    """
    Validate the file naming format.
    
    Parameters
    ----------
    file_naming_format : list, optional
        The order of components in the filename. 
        Options: ["data_category", "cohort_y11_ay", "cohort_yg_ay", "year_group"].
        Default: ["cohort_y11_ay", "data_category" , "cohort_yg_ay", "year_group"].
    
    Return
    -------
    list
      
    """
     # Ensure all components in file_naming_format are lowercase
    file_naming_format = [component.lower() for component in file_naming_format]
    
    # -----------------------
    # Validation steps:
    # Validate file_naming_format: Ensure only valid components are included
    valid_components = {"data_category", "cohort_y11_ay", "cohort_yg_ay", "year_group"}
    valid_format_example = (
        '["cohort_y11_ay", "data_category", "year_group"] '
        'or ["data_category", "cohort_y11_ay", "cohort_yg_ay", "year_group"]'
    )
    
    # Validate file_naming_format length (must be 3 or 4)
    if len(file_naming_format) not in [3, 4]:
        raise ValueError(
            f"Invalid file_naming_format length. It must contain exactly 3 or 4 components. "
            f"Valid components are: {valid_components}."
            f"For example: {valid_format_example}."
        )
    
    # Validate file_naming_format components
    invalid_components = set(file_naming_format) - valid_components
    if invalid_components:
        raise ValueError(
            f"Invalid components in file_naming_format: {invalid_components}. "
            f"Valid components are: {valid_components}. "
            f"For example: {valid_format_example}."
        )
    
     # Ensure required components are present in file_naming_format
    if "data_category" not in file_naming_format or "cohort_y11_ay" not in file_naming_format:
        raise ValueError(
            "file_naming_format must include 'data_category' and 'cohort_y11_ay'."
        )
    # -----------------------
    
     # Handle 4-component case
    if len(file_naming_format) == 4:
        # If both year cohort and year_group are present, prioritize the earlier component and
        # remove the one with the larger index
        if "cohort_yg_ay" in file_naming_format and "year_group" in file_naming_format:
            cohort_index = file_naming_format.index("cohort_yg_ay")
            year_group_index = file_naming_format.index("year_group")

            if cohort_index < year_group_index:
                file_naming_format.pop(year_group_index)  # Remove year_group
            else:
                file_naming_format.pop(cohort_index)  # Remove cohort
    
    return file_naming_format

def extract_y11_cohort(file_name: str, date_format: str) -> str:
    # Extract cohort information from the file name based on the specified date format
    try:
        if date_format == "MMM YY": # e.g., Mar 18
            match = re.search(
                r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec) (\d{2})', file_name)
            if match:
                month = match.group(1)
                year = '20' + match.group(2)
                return f'{year} {month}'
        elif date_format == "YYYY_YY": # e.g., 2018_19
            match = re.search(r'(\b\d{4}_\d{2}\b)', file_name)
            if match:
                date_str = match.group(1)
                start_year, end_year_suffix = date_str.split('_')
                end_year = str(int(start_year[:2] + end_year_suffix))
                return f"{start_year} - {end_year}"
        elif date_format == "YYYY": # e.g., 1819, stands for 2018 - 2019
            match = re.search(r'(\b\d{4}\b)', file_name)
            if match:
                date_str = match.group(1)
                start_year = "20" + date_str[:2]
                end_year = "20" + date_str[2:]
                return f"{start_year} - {end_year}"

        # Otherwise Try to find a year in brackets, use that as the end year, e.g., (Y11 2018)
        print("Trying to find a year in brackets to use as the end year of Y11...")
        match = re.search(r'\([^)]*(\d{4})[^)]*\)', file_name)
        if match:
            year = match.group(1)
            return f"{str(int(year) - 1)} - {year}"
        
    except (ValueError, AttributeError) as e:
        print(f"Error parsing date from file '{file_name}': {e}")
        return None
    return None

def extract_year_group(sheet_name: str) -> tuple:
    # Extract year group information from the sheet name, the sheet should include yearly data for a specific Y11 cohort.
    sheet_name = sheet_name.strip()

    # Rule 1: e.g. for 'Y11 201819 Census', it matches 'Y11 201819'
    match = re.match(r'\b(Y\d+) (\d{6})\b(?!.*\d{2}\b)', sheet_name)
    if match:
        if '_' not in sheet_name:
            year_group = match.group(1)
            year_group = int(year_group.replace('Y', ''))
            year_suffix = match.group(2)
            start_year = int(year_suffix[:4])
            end_year = int(year_suffix[4:]) + 2000
            cohort_yg_ay = f'{year_group} {start_year} - {end_year}'
            return year_group, cohort_yg_ay

    # Rule 2: e.g. for 'AS Y11 1819 at 011119', it matches 'Y11 1819'
    # BUG: Will be problematic for 'Y10 2024 Oct 23 Census'
    match = re.search(r'\b(Year \d+|Y\d+) (\d{2}\d{2})\b', sheet_name)
    if match:
        year_group = match.group(1).replace('Year ', 'Y')
        year_group = int(year_group.replace('Y', ''))
        year_suffix = match.group(2)
        start_year = int(year_suffix[:2]) + 2000
        end_year = int(year_suffix[2:]) + 2000
        cohort_yg_ay = f'{year_group} {start_year} - {end_year}'
        return year_group, cohort_yg_ay

    # Rule 3: e.g. for 'Y11 202122 Attendance 17_18', it matches 'Y11 17_18'
    match = re.search(r'\b(Y\d+).*?\b(\d{2})_(\d{2})\b', sheet_name)
    if match:
        year_group = match.group(1)
        year_group = int(year_group.replace('Y', ''))
        start_year = int(match.group(2)) + 2000
        end_year = int(match.group(3)) + 2000
        cohort_yg_ay = f'{year_group} {start_year} - {end_year}'
        return year_group, cohort_yg_ay

    # Rule 4: e.g. for 'Y8 201819 Jan 16 Census', it matches 'Y8 201819 Jan 16'
    match = re.search(r'\b(Y\d+) (\d{6}).*?\b(Jan)\b.*?(\d{2})', sheet_name)
    if match:
        year_group = match.group(1)
        year_group = int(year_group.replace('Y', ''))
        end_year = int(match.group(4)) + 2000
        start_year = end_year - 1
        cohort_yg_ay = f'{year_group} {start_year} - {end_year}'
        return year_group, cohort_yg_ay

    # Rule 5: e.g. for 'Y10 in Y7 (2020_21)', it matches 'Y7 (2020_21)'
    match = re.search(r'Y(\d+)\s*\((\d{4})_(\d{2})\)', sheet_name)
    if match:
        year_group = int(match.group(1))     
        start_year = int(match.group(2))      
        end_year = int(match.group(3)) + 2000
        cohort_yg_ay = f'{year_group} {start_year} - {end_year}'
        return year_group, cohort_yg_ay

    # Rule 6: e.g. for '2024 Y11 - Excl from Y9 (21_22)', it matches 'Y9 (21_22)'
    match = re.search(r'Y(\d+)\s*\((\d{2})_(\d{2})\)', sheet_name)
    if match:
        year_group = int(match.group(1))     
        start_year = int(match.group(2)) + 2000
        end_year = int(match.group(3)) + 2000
        cohort_yg_ay = f'{year_group} {start_year} - {end_year}'
        return year_group, cohort_yg_ay

    # Rule 7: e.g. for '2018_19 (Y7)', it matches '2018_19 (Y7)'
    match = re.search(r'(\d{4})_(\d{2})\s*\(Y(\d+)\)', sheet_name)
    if match:
        start_year = int(match.group(1))
        end_year = int(match.group(2)) + 2000
        year_group = int(match.group(3))
        cohort_yg_ay = f'{year_group} {start_year} - {end_year}'
        return year_group, cohort_yg_ay

    # Rule 8: e.g. for 'Y11 in Y10 (Jan 23 Census)', it matches 'Y10 (Jan 23 Census)'
    match = re.search(r'Y(\d+)\s*\(([^)]*)\)', sheet_name)
    if match:
        year_group = int(match.group(1))
        bracket_text = match.group(2)
        two_digit_numbers = re.findall(r'\b(\d{2})\b', bracket_text) # only 2 digits within the bracket
        if len(two_digit_numbers) == 1:
            end_year_suffix = int(two_digit_numbers[0])
            end_year = end_year_suffix + 2000
            start_year = end_year - 1
            cohort_yg_ay = f'{year_group} {start_year} - {end_year}'
            return year_group, cohort_yg_ay
    
    # Rule 9: e.g., for 'All Susp and Perm Excl 18-19', it matches '18-19'
    match = re.search(r'^[^\d]*?(\d{2})-(\d{2})\b', sheet_name)
    if match:
        start_year = int(match.group(1)) + 2000
        end_year = int(match.group(2)) + 2000
        year_group = 11 # use 11 for default year_group value
        cohort_yg_ay = f'{year_group} {start_year} - {end_year}'
        return year_group, cohort_yg_ay
    
    # Rule 10: e.g., for 'When Y10 in 202223' it matches 'Y10 202223'
    yg_matches = re.findall(r'\bY\d+\b', sheet_name)
    year_matches = re.findall(r'\b\d{6}\b', sheet_name)

    if len(yg_matches) == 1 and len(year_matches) == 1:
        year_group = int(yg_matches[0].replace('Y', ''))
        start_year = int(year_matches[0][:4])
        end_year = int(year_matches[0][4:]) + 2000
        cohort_yg_ay = f'{year_group} {start_year} - {end_year}'
        return year_group, cohort_yg_ay

    return None, None

def data_upload_report(file_metadata: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    """
    Generate a summary report of existing, missing, and combined data categories for each year cohort.
    
    This function examines the provided file_metadata DataFrame and summarises both the 
    existing data categories and missing data categories for each year cohort and year group.
    Additionally, it checks if any year group is missing for each Y11 cohort.

    Parameters
    ----------
    file_metadata : pd.DataFrame
        A DataFrame containing file information with columns defined in FileMetadata such as 
        - FileMetadata.COHORT_Y11_AY, 
        - FileMetadata.COHORT_YG_AY, 
        - FileMetadata.CATEGORY, 
        - FileMetadata.YEAR_GRP, 
        - FileMetadata.FILE_NAME, 
        - FileMetadata.SHEET_NAME.

    Returns
    -------
    report_dict : dict
        A dictionary containing the following DataFrames:
        - 'existing_categories': Summarising the existing data categories for each year cohort and year group.
        - 'missing_categories': Summarising the missing data categories for each year cohort and year group.
        - 'combined_categories': Summarising both existing and missing data categories for each year cohort and year group.
        - 'missing_year_groups': Summarising the missing year groups for each Y11 Cohort.
    """
    
    # Required columns for the function to operate correctly
    required_columns = {
        FileMetadata.CATEGORY,
        FileMetadata.COHORT_Y11_AY,
        FileMetadata.COHORT_YG_AY,
        FileMetadata.YEAR_GRP,
        FileMetadata.FILE_NAME,
        FileMetadata.SHEET_NAME,
    }


    # Validate if the necessary columns are present
    if not required_columns.issubset(file_metadata.columns):
        missing_columns = required_columns - set(file_metadata.columns)
        raise ValueError(
            f"The input DataFrame is missing the required columns: {', '.join(missing_columns)}. "
            f"Expected columns are: {', '.join(required_columns)}"
        )
    
    existing_categories_summary = []
    missing_categories_summary = []
    missing_year_groups_summary = []
    
    # Determine all possible year groups across all Y11 Cohorts
    all_possible_year_groups = sorted(
        file_metadata[FileMetadata.YEAR_GRP].dropna().unique().astype(int)
    )
    
    # Group the file_metadata DataFrame by (year_group, cohort_yg_ay)
    # - `year_group` represents the school year in which students were in a particular year group 
    #   (e.g., Year 7 in 2014-2015, Year 7 in 2015-2016).
    # - `cohort_yg_ay` refers to the specific academic year when the students were in that year group.
    for (year_group, cohort_yg_ay), cohort_group in file_metadata.groupby(
        [FileMetadata.YEAR_GRP, FileMetadata.COHORT_YG_AY]
    ):
        assert year_group.is_integer() # Year Group must be an integer
        year_group = int(year_group)  # Ensure Year Group is an integer
        
        # Determine the possible data categories for this specific year group
        all_data_categories = sorted(
            set(
                file_metadata[
                    file_metadata[FileMetadata.YEAR_GRP] == year_group
                ][FileMetadata.CATEGORY].dropna().unique()
            )
        )
        
        # Remove `FileMetadata.EXCLUDE` if present and filter out categories starting with '~'
        # (which indicates unfinished data that needs further special handling)
        all_data_categories = [
            cat for cat in all_data_categories 
            if cat != FileMetadata.EXCLUDE and not cat.startswith("~")
        ]

        # Extract available data categories, ensuring they are non-null and unique
        available_data_categories = sorted({
            cat for cat in cohort_group[FileMetadata.CATEGORY] if pd.notna(cat)
        })

        # Identify missing data categories by finding the difference from all_data_categories
        missing_data_categories = sorted(set(all_data_categories) - set(available_data_categories))
        
        # Add corresponding file name in "{File Name} - {Sheet Name}" format
        file_and_sheet_names = (
            cohort_group.apply(
                lambda row: f"{row[FileMetadata.FILE_NAME]} - {row[FileMetadata.SHEET_NAME]}",
                axis=1
            )
            .unique()
        )
        
        file_and_sheet_names_str = ', '.join(file_and_sheet_names)
        
        y11_cohort_unique = cohort_group[FileMetadata.COHORT_Y11_AY].unique()

        # Ensure that there is only one unique value for Y11 Cohort
        if len(y11_cohort_unique) == 1:
            y11_cohort_value = y11_cohort_unique[0]
        else:
            raise ValueError(f"More than one Y11 Cohort ({y11_cohort_unique}) found for Year Cohort '{cohort_yg_ay}'.")

        existing_categories_summary.append({
            FileMetadata.COHORT_Y11_AY: y11_cohort_value, 
            FileMetadata.YEAR_GRP: year_group,
            FileMetadata.COHORT_YG_AY: cohort_yg_ay,
            'Existing Categories': ', '.join(available_data_categories),
            'File Name - Sheet Name': file_and_sheet_names_str
        })
        
        if missing_data_categories:
            missing_categories_summary.append({
                FileMetadata.COHORT_Y11_AY: y11_cohort_value,
                FileMetadata.YEAR_GRP: int(year_group),
                FileMetadata.COHORT_YG_AY: cohort_yg_ay,
                'Missing Categories': ', '.join(missing_data_categories),
                'File Name - Sheet Name': file_and_sheet_names_str
            })
    
    # Check for missing year groups within each Y11 Cohort
    for cohort_y11_ay, cohort_group in file_metadata.groupby(FileMetadata.COHORT_Y11_AY):
        if cohort_group[FileMetadata.COHORT_YG_AY].isna().all():
            continue # NCCIS data won't be considered.

        available_year_groups = sorted(
            cohort_group[FileMetadata.YEAR_GRP].dropna().unique().astype(int)
        )
        
        missing_year_groups = sorted(
            set(all_possible_year_groups) - set(available_year_groups)
        )
        
        # If there are missing year groups, add each one as a separate row
        for year_group in missing_year_groups:
            year_group = int(year_group)
            missing_year_groups_summary.append({
                FileMetadata.COHORT_Y11_AY: cohort_y11_ay,
                FileMetadata.YEAR_GRP: year_group,
                FileMetadata.COHORT_YG_AY: (
                    f"Y{year_group} {int(cohort_y11_ay.split()[0]) - (11 - year_group)} "
                    f"- {int(cohort_y11_ay.split()[-1]) - (11 - year_group)}"
                ),
                # 'File Name - Sheet Name': '',
                # 'Existing Categories': '',
                'Missing Categories': 'ks2' if year_group == 6 else 'All data missing for this year group',
            })
    
    required_cols_report = [
        FileMetadata.COHORT_Y11_AY,
        FileMetadata.YEAR_GRP,
        FileMetadata.COHORT_YG_AY,
        "Existing Categories",
        "File Name - Sheet Name",
    ]

    # Convert summaries into DataFrames, sort them, and handle empty cases
    existing_categories_df = (
        pd.DataFrame(existing_categories_summary)
        .sort_values(by=[FileMetadata.COHORT_Y11_AY, FileMetadata.YEAR_GRP])
        if existing_categories_summary
        else pd.DataFrame(columns=required_cols_report)
    )

    missing_categories_df = (
        pd.DataFrame(missing_categories_summary)
        .sort_values(by=[FileMetadata.COHORT_Y11_AY, FileMetadata.YEAR_GRP])
        if missing_categories_summary
        else pd.DataFrame(columns=required_cols_report)
    )

    missing_year_groups_df = (
        pd.DataFrame(missing_year_groups_summary)
        .sort_values(by=[FileMetadata.COHORT_Y11_AY, FileMetadata.YEAR_GRP])
        if missing_year_groups_summary
        else pd.DataFrame(columns=required_cols_report)
    )
    
    # merge existing and missing categories
    report_df = pd.merge(
        existing_categories_df,
        missing_categories_df,
        on=[FileMetadata.COHORT_Y11_AY, FileMetadata.YEAR_GRP,
            FileMetadata.COHORT_YG_AY, 'File Name - Sheet Name'],
        how='outer'
    ).fillna('')

    # concatenate the missing year groups DataFrame
    report_df = pd.concat(
        [report_df, missing_year_groups_df],
        ignore_index=True,
        sort=False
    ).fillna('').sort_values(by=[FileMetadata.COHORT_Y11_AY, FileMetadata.YEAR_GRP])
    
    missing_report_df = pd.concat(
        [missing_categories_df, missing_year_groups_df],
        ignore_index=True,
        sort=False
    ).fillna('').sort_values(by=[FileMetadata.COHORT_Y11_AY, FileMetadata.YEAR_GRP])
    
    # Count the number of unique Year 11 cohorts that are formatted either as "YYYY - YYYY" or "YYYY-YYYY"
    num_cohorts = (
        file_metadata[FileMetadata.COHORT_Y11_AY]
        .dropna()
        .loc[lambda x: x.str.match(r'^\d{4}\s?-\s?\d{4}$')]
        .nunique()
    )
    
    # Count the range of Year 11 cohorts
    cohort_range = (
        f"{file_metadata[FileMetadata.COHORT_Y11_AY].dropna().min()} "
        f"to {file_metadata[FileMetadata.COHORT_Y11_AY].dropna().max()}"
    )
    
    # # Count the number of unique Year Groups (not entirely accurate as NCCIS data includes students up to age 25)
    # num_year_groups = file_metadata[FileMetadata.YEAR_GRP].nunique()
    # # Count the range of unique Year Groups
    # year_group_range = f"{file_metadata[FileMetadata.YEAR_GRP].min().astype(int)} to {file_metadata[FileMetadata.YEAR_GRP].max().astype(int)}"
    
    # List the uploaded data types
    data_types_needed = file_metadata[FileMetadata.CATEGORY].dropna()
    exclude_str = re.escape(FileMetadata.EXCLUDE)  
    data_types_needed = data_types_needed[
        ~data_types_needed.str.contains(f'^~|{exclude_str}', regex=True)
    ].unique().tolist()
    
    # Count the number of Year 11 cohorts that have missing data.
    missing_data_cohorts = missing_report_df[FileMetadata.COHORT_Y11_AY].nunique()
    
    summary_data = {
        'Summary': [
            'Number of Year 11 Cohorts',
            'Year 11 Cohorts Range',
            # 'Number of Year Groups',
            # 'Year Group Range',
            'Uploaded Data Types',
            'Y11 Cohorts with Missing Data'
        ],
        'Details': [
            num_cohorts,
            cohort_range,
            # num_year_groups,
            # year_group_range,
            ', '.join(data_types_needed),
            missing_data_cohorts
        ]
    }
    summary_df = pd.DataFrame(summary_data)

    
    report_dict = {
        'existing_categories': existing_categories_df,
        'missing_categories': missing_categories_df,
        'missing_year_groups': missing_year_groups_df,
        'missing_report': missing_report_df,
        'full_report': report_df,
        'summary': summary_df,
    }

    return report_dict

def prefix_columns(
    df: pd.DataFrame, 
    prefix: str,
    stud_id_col: str = STUD_ID_COL, 
) -> pd.DataFrame:
    """
    Add a prefix to columns in a DataFrame, except for the student ID column and already prefixed columns.

    Parameters
    ----------
    df : pd.DataFrame
        The DataFrame whose columns need to be prefixed.

    stud_id_col : str
        The name of the student ID column that should not be prefixed.

    prefix : str
        The prefix to add to columns.

    Returns
    -------
    pd.DataFrame
        The DataFrame with updated column names.
    """
    df.columns = [
        col if col == stud_id_col or col.startswith(prefix) or col.startswith('_')
        else f"{prefix}_{col}"
        for col in df.columns
    ]
    return df

def add_prefix(
    colname: str,
    prefix: str
) -> str:
    new_colname = f"{prefix}_{colname}"
    return new_colname

def remove_prefix(
    colname: str, 
    valid_prefixes: set = None
) -> str:
    """Remove the prefix from a column name if it matches a valid prefix."""
    
    if valid_prefixes is None:
        valid_prefixes = set(DATA_CATEGORIES)
        
    for prefix in valid_prefixes:
        if colname.startswith(f"{prefix}_"):
            return colname[len(prefix) + 1:]  # Remove prefix + underscore
    return colname 

def resolve_column_name(
    colname: Union[str, List[str]], 
    df_columns: Union[Set[str], List[str]], 
    valid_prefixes: Union[str, Set[str]]
) -> Union[str, List[str], None]:
    """
    Resolve column names by checking whether they exist in the DataFrame columns
    with or without a prefix.

    Parameters
    ----------
    colname : str or list of str
        The column name(s) to check (which may or may not have a prefix).
    df_columns : set or list
        The set or list of column names in the DataFrame.
    valid_prefixes : str or set
        The valid prefix or a set of valid prefixes used in column names.

    Returns
    -------
    str or list of str or None
        - If input is a string, returns the resolved column name (or None if not found).
        - If input is a list, returns a list of resolved column names (or None for missing columns).

    Examples
    --------
    df_columns = {"nccis_academic_age", "confirmed_date", "stud_id"}

    resolve_column_name("academic_age", df_columns, "nccis")  # Returns "nccis_academic_age"
    resolve_column_name(["academic_age", "confirmed_date"], df_columns, {"nccis"})  # Returns ["nccis_academic_age", "confirmed_date"]
    resolve_column_name("missing_col", df_columns, "nccis")  # Returns None
    """

    # Ensure df_columns is a set for faster lookups
    if isinstance(df_columns, list):
        df_columns = set(df_columns)

    # Ensure valid_prefixes is a set
    if isinstance(valid_prefixes, str):
        valid_prefixes = {valid_prefixes}

    # If colname is a list, resolve each column separately
    if isinstance(colname, list):
        return [resolve_column_name(single_col, df_columns, valid_prefixes) for single_col in colname]

    # If colname is a single string, apply the resolution logic
    # Check if the column exists as is
    if colname in df_columns:
        return colname

    # Try adding a prefix and check
    for prefix in valid_prefixes:
        prefixed_col = add_prefix(colname, prefix)
        if prefixed_col in df_columns:
            return prefixed_col

    # Try removing a prefix and check
    unprefixed_col = remove_prefix(colname, valid_prefixes)
    if unprefixed_col in df_columns:
        return unprefixed_col

    # If no match is found
    return None

def compute_data_hash(df: pd.DataFrame) -> str:
    """
    Compute the hash of a DataFrame to identify duplicate data.
    
    Parameters
    ----------
    df : pd.DataFrame
        The DataFrame to hash.
    
    Returns
    -------
    str
        The hash value of the DataFrame.
    """
    data_bytes = df.to_csv(index=False).encode('utf-8')
    return hashlib.md5(data_bytes).hexdigest()

def count_file_rows(file_path, sheet_name=None, has_header=True):
    ext = os.path.splitext(file_path)[1].lower()

    if ext == '.csv':
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            row_count = sum(1 for _ in reader)
            return row_count - 1 if has_header else row_count

    elif ext in ['.xlsx', '.xls']:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheet = wb[sheet_name or wb.sheetnames[0]]
        row_count = sheet.max_row
        return row_count - 1 if has_header else row_count

    else:
        raise ValueError(f"Unsupported file type: {ext}")

