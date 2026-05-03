import pandas as pd
import re
import logging
from pathlib import Path
from typing import List, Dict, Union
from tabulate import tabulate
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

from ...utils.constants import (
    FileMetadata,
    ColumnMetadata,
    ColumnSchema,
    NCCIS
)

from ...utils.misc import (
    styled_print, 
    parse_yaml, 
    get_files_in_folder, 
    list_sheets_in_excel, 
    get_differences
)

from ...utils.logger_setup import (
    get_logger, 
    log_with_border, 
)

from .._utils import (
    extract_y11_cohort,
    extract_year_group,
    data_upload_report,
    count_file_rows,
    remove_prefix
)

from ._metadata_utils import (
    create_file_metadata_entry,
    auto_curate_cohort_metadata,
    compare_and_backup
)  

logger = get_logger("data_processor")
# logger = logging.getLogger(__name__)  


def extract_file_meta_from_dir(
    folder_path: Union[str, Path],
    file_naming_format: list,
) -> pd.DataFrame:
    """
    Generate a file metadata DataFrame by extracting details from standardised file names.

    Parameters
    ----------
    folder_path : Union[str, Path]
        The path to the folder containing the standardised files.
    
    file_naming_format : list
        The order of components in the filename. 
        Options: ["data_category", "cohort_y11_ay", "cohort_yg_ay", "year_group"].
        Default: ["cohort_y11_ay", "data_category" , "year_group"].

    Returns
    -------
    pd.DataFrame
        A DataFrame containing file information extracted from file names.
    """
    
    styled_print(
        f"Starting to generate file metadata from folder {folder_path}. "
        "Please double-check the generated file metadata, as it may be inaccurate.",
        colour="yellow"
    )

    folder_path = Path(folder_path)
    file_metadata_list = []
    
    for file_path in sorted(folder_path.glob('*.*')):
        if file_path.name.startswith('.'):
            continue  # Skip hidden files

        try:
            name_parts = file_path.stem.split('_')
            
            data_category = name_parts[file_naming_format.index('data_category')]
            cohort_y11_ay = name_parts[file_naming_format.index('cohort_y11_ay')]
            
            # Conditionally extract 'year_group' and 'cohort_yg_ay' if they are present in the naming format
            year_group = (
                name_parts[file_naming_format.index('year_group')]
                if 'year_group' in file_naming_format and len(name_parts) > file_naming_format.index('year_group')
                else None
            )
            cohort_yg_ay = (
                name_parts[file_naming_format.index('cohort_yg_ay')]
                if 'cohort_yg_ay' in file_naming_format and len(name_parts) > file_naming_format.index('cohort_yg_ay')
                else None
            )
            
            if cohort_yg_ay and cohort_yg_ay.lower().startswith('y'):
                year_group = int(year_group.split('-')[0][1:])
         
            if year_group and not cohort_yg_ay:
                # Fill the cohort_yg_ay information if it is missing
                _, cohort_yg_ay = auto_curate_cohort_metadata(year_group, cohort_yg_ay, cohort_y11_ay)

            # Append to file_metadata_list
            file_metadata_list.append({
                FileMetadata.CATEGORY: data_category,
                FileMetadata.STD_FILE_NAME: file_path.name,
                FileMetadata.COHORT_Y11_AY: cohort_y11_ay,
                FileMetadata.YEAR_GRP: int(year_group) if year_group else pd.NA,
                FileMetadata.COHORT_YG_AY: cohort_yg_ay if cohort_yg_ay else pd.NA,
                FileMetadata.ROW_COUNT: count_file_rows(file_path)
            })

        except Exception as e:
            logger.warning(
                f"Failed to parse file name: {file_path.name} - Error: {e}"
            )

    # Convert the list to a DataFrame
    df = pd.DataFrame(file_metadata_list).sort_values(
        by=[
            FileMetadata.CATEGORY, 
            FileMetadata.COHORT_Y11_AY, 
            FileMetadata.STD_FILE_NAME
            ]
        )

    styled_print("File metadata generated successfully.", colour="magenta")

    return df

# TODO: use extract_file_metadata_from_folder() if yaml file not provided
def extract_file_metadata(
    folder_path: Union[str, Path],
    yaml_file_path: Union[str, Path],
    output_path: Union[str, Path],
    valid_cat: list,
    overwrite: bool = False,
    display_upload_status: bool = True,
    has_processed: bool = False
) -> pd.DataFrame:
    """
    Generate a file metadata DataFrame with detailed information about each
    file and sheet

    Parameters
    ----------
    folder_path : Union[str, Path]
        The path to the folder containing the files to be processed.
    
    yaml_file_path : Union[str, Path]
        The path to the YAML configuration file.
    
    output_path : Union[str, Path]
        The path where the output file metadata DataFrame should be saved.
    
    valid_cat : list
        A list of valid data categories to consider.
    
    display_upload_status : bool, optional
        Whether to display the upload status summary. Defaults to True.
    
    has_processed : bool, optional
        Whether the files have been processed already. Defaults to False.
    
    overwrite : bool, optional
        Flag indicating whether to overwrite the existing file metadata file, by default False.

    Returns
    -------
    pd.DataFrame
        A DataFrame containing detailed information about each file and sheet.

    Notes
    -----
    This function processes files in the specified folder according to the rules defined
    in the YAML configuration file. It generates a DataFrame with detailed information 
    about each file and sheet, including:

    - Data Category: Assigned based on the existence of keywords in the file name or sheet name.
    - Y11 Cohort: The cohort of students in Year 11.
    - Sheet Name: The name of the sheet within the file.
    - Year Group: The school year the data was recorded for (e.g., 9 for Year 9).
    - Year Cohort: The cohort year the data pertains to (e.g., 2016-2017 for the cohort's Year 9 data).
    - Uncurated Year Group and Cohort: The initial, uncurated values of Year
      Group and Cohort based on corresponding file/sheet name.
  
    Example of the YAML configuration file structure:
    ```yaml
    attendance: # data category (without underscore, used for logical groupings for files)
      description: Attendance data
      is_meta: False
      cohort_date_format: YYYY_YY # date format for Y11 cohort, e.g. 2025_26
      keywords: 
        - attend
        - attendance
      metadata:
        sources:
          - filename: School data metadata (census, attendance, exclusions).xlsx
            sheetname: Attendance
            colname: Field
            description: Info
            value: Values
    ```

    If an existing file metadata file is found and overwrite is set to False, this function 
    will append new information to the existing file metadata rather than overwriting it.
    """

    # log_with_border(logger, "Starting extracting file metadata...")

    yaml_data = parse_yaml(yaml_file_path)
    file_metadata_list = []
    files = get_files_in_folder(folder_path)
    logger.info(f"Number of files found in {folder_path}: {len(files)}")

    # If output file exists and overwrite is False, load existing file_metadata file
    if output_path.exists() and not overwrite:
        existing_file_metadata = pd.read_excel(output_path)

        existing_file_metadata[FileMetadata.SHEET_NAME] = existing_file_metadata[FileMetadata.SHEET_NAME].apply(
            lambda x: None if pd.isna(x) else x
        )

        # track existing files and sheets
        existing_files_sheets = set(
            existing_file_metadata.apply(
                lambda row: (row[FileMetadata.FILE_NAME], row[FileMetadata.SHEET_NAME]), axis=1
            )
        )

        file_metadata_list = existing_file_metadata.to_dict("records")
    else:
        existing_files_sheets = set()

    new_entries_added = False  # Track if new entries are added

    # Process each file
    for file_path in files:
        file_name = file_path.name
        file_name_lower = file_name.replace('ANON', '').lower()

        if file_name.endswith('.xlsx'):
            all_sheets = list_sheets_in_excel(file_path)
        else:
            all_sheets = [None]

        for sheet_name in all_sheets:
            sheet_name_lower = sheet_name.lower() if sheet_name else ''

            # Skip processing if file and sheet are already in the existing file_metadata
            if (file_name, sheet_name) in existing_files_sheets:
                # logger.info(f"Skipping already processed file and sheet: {file_name}, {sheet_name}")
                continue

            data_category_matched = None

            # check if the file and sheet should be excluded
            exclude_config = yaml_data.get(FileMetadata.EXCLUDE, {})
            
            # If exclude_config is empty, try to ignore case by checking for lowercase keys
            if not exclude_config:
                # Try to find a key that matches 'exclude' in a case-insensitive way
                for key in yaml_data.keys():
                    if key.lower() == FileMetadata.EXCLUDE.lower():
                        exclude_config = yaml_data[key]
                        break

            for exclude_rule in exclude_config.get('files', []):
                name_contains = exclude_rule.get('name_contains', '').lower()
                sheets_contain = exclude_rule.get('sheets_contain', [])

                if name_contains in file_name_lower:
                    if not sheets_contain or any(sc.lower() in sheet_name_lower for sc in sheets_contain):
                        data_category_matched = FileMetadata.EXCLUDE
                        break

            if data_category_matched == FileMetadata.EXCLUDE:
                file_metadata_list.append(create_file_metadata_entry(
                    data_category=data_category_matched,
                    file_name=file_name,
                    sheet_name=sheet_name,
                ))
                
                existing_files_sheets.add((file_name, sheet_name))
                
                new_entries_added = True
                
                logger.warning(
                    f"File '{file_name}', sheet '{sheet_name}' is excluded according to EXCLUDE rules."
                )
                continue  # Skip to next sheet

            # check other data categories
            for data_category, config in yaml_data.items():
                if data_category == FileMetadata.EXCLUDE or config.get('is_meta', False):
                    logger.info(f"Excluding category: {data_category}")
                    continue

                keywords = config.get('keywords', [])
                if isinstance(keywords, str):
                    keywords = [keywords]
                keywords_lower = [kw.lower() for kw in keywords]

                # Check if any keyword is in file_name or sheet_name
                if any(kw in file_name_lower for kw in keywords_lower) or any(kw in sheet_name_lower for kw in keywords_lower):
                    # Check filter_criteria if any
                    filter_criteria = config.get('filter_criteria', {}).get('contains', [])
                    if isinstance(filter_criteria, str):
                        filter_criteria = [filter_criteria]
                    filter_criteria_lower = [fc.lower() for fc in filter_criteria]

                    if filter_criteria_lower:
                        if not any(fc in file_name_lower for fc in filter_criteria_lower):
                            # Does not match filter criteria, skip to next data category
                            continue

                    # We have a matching data category
                    data_category_matched = data_category
                    cohort_date_format = config.get('cohort_date_format', '')
                    break  # Stop after first match

            if data_category_matched:
                
                # Extract the academic year in which the cohort reached Year 11 from file name
                cohort_y11_ay = extract_y11_cohort(file_name, cohort_date_format)

                # Extract year group and corresponding academic year
                year_group, cohort_yg_ay = extract_year_group(
                    sheet_name) if sheet_name is not None else (None, None)
                
                uncurated_years, uncurated_cohort_yg_ay = year_group, cohort_yg_ay

                if data_category_matched.lower() == 'ks2':
                    year_group = 6
                    uncurated_years = 6
                    uncurated_cohort_yg_ay = f"{year_group} {cohort_y11_ay}"
                elif data_category_matched.lower() == 'ks4':
                    year_group = 11
                    uncurated_years = 11
                    uncurated_cohort_yg_ay = f"{year_group} {cohort_y11_ay}"

                # Auto-curation
                year_group, cohort_yg_ay = auto_curate_cohort_metadata(
                    year_group, uncurated_cohort_yg_ay, cohort_y11_ay)
                
                review_flag = False
                
                if pd.isna(cohort_y11_ay):
                    review_flag = True
                elif data_category not in [FileMetadata.EXCLUDE, NCCIS.MAR_VER, NCCIS.SEP_VER]:
                    if pd.isna(year_group) or pd.isna(cohort_yg_ay):
                        review_flag = True

                file_metadata_list.append(create_file_metadata_entry(
                    data_category=data_category_matched,
                    file_name=file_name,
                    cohort_y11_ay=cohort_y11_ay,
                    sheet_name=sheet_name,
                    year_group=year_group,
                    cohort_yg_ay=cohort_yg_ay,
                    uncurated_year_group=uncurated_years,
                    uncurated_cohort_yg_ay=uncurated_cohort_yg_ay,
                    row_counts=count_file_rows(file_path),
                    needs_review='YES' if review_flag else None
                ))

                existing_files_sheets.add((file_name, sheet_name))
                new_entries_added = True
                if review_flag:  # If need to manually fill y11 cohort
                    logger.warning(
                        f"Appending {(file_name, sheet_name)} to {output_path}, but manual check is needed "
                        "as only data category and/or Y11 academic year could be extracted from the metadata."
                    )
                else: 
                    logger.info(f'Appending {file_name, sheet_name} to {output_path}')
            else:
                # No matching data category     
                file_metadata_list.append(
                    create_file_metadata_entry(
                        file_name=file_name,
                        sheet_name=sheet_name,
                        row_counts=count_file_rows(file_path),
                        needs_review="YES"
                    )
                )
                existing_files_sheets.add((file_name, sheet_name))
                new_entries_added = True
                logger.warning(
                    f"Appending {(file_name, sheet_name)} to {output_path}, but needs further manual check "
                    "as no data category or metadata can be extracted..."
                )

    df = (
        pd.DataFrame(file_metadata_list)
        .sort_values(
            by=[
                FileMetadata.CATEGORY,
                FileMetadata.COHORT_Y11_AY,
                FileMetadata.FILE_NAME,
                FileMetadata.YEAR_GRP,
                FileMetadata.SHEET_NAME,
            ]
        )
    )

    # Then check if manually filled information (year group, y11 cohort, year cohort)
    updates_made = False
    df[FileMetadata.NEEDS_REVIEW] = df[FileMetadata.NEEDS_REVIEW].fillna('')
    for index, row in df.iterrows():
        if pd.isna(row[FileMetadata.RAW_YEAR_GRP]) or row[FileMetadata.NEEDS_REVIEW].lower() == 'yes':
            data_category = row[FileMetadata.CATEGORY]
            original_year_group = row[FileMetadata.YEAR_GRP]
            cohort_y11_ay = row[FileMetadata.COHORT_Y11_AY]
            original_uncurated_cohort = row[FileMetadata.COHORT_YG_AY]

            if pd.notna(data_category) and data_category != FileMetadata.EXCLUDE:
                if data_category.lower() == 'ks2': # can be improved by using custom configs from data_configs.yaml
                    year_group = 6
                    uncurated_years = 6
                    uncurated_cohort_yg_ay = f"{year_group} {cohort_y11_ay}"
                elif data_category.lower() == 'ks4':
                    year_group = 11
                    uncurated_years = 11
                    uncurated_cohort_yg_ay = f"{year_group} {cohort_y11_ay}"
                else:
                    year_group = original_year_group
                    uncurated_cohort_yg_ay = original_uncurated_cohort

                curated_year_group, curated_cohort = auto_curate_cohort_metadata(
                    year_group, uncurated_cohort_yg_ay, cohort_y11_ay)

                # Check if the year group needs to be updated based on manually filled data.
                # The update occurs if both the curated year group and the manually filled year group are not empty, and the curated year group is different from the original year group.
                if not (pd.isna(curated_year_group) and pd.isna(original_year_group)) and curated_year_group != original_year_group:
                    df.at[index, FileMetadata.YEAR_GRP] = curated_year_group
                    logger.info(
                        f"The Year Group of '{row[FileMetadata.FILE_NAME]} - {row[FileMetadata.SHEET_NAME]}' "
                        f"has been updated from {original_year_group} to {curated_year_group}"
                    )
                    updates_made = True

                # Check if need to update manually filled FileMetadata.COHORT_YG_AY column
                if not (pd.isna(curated_cohort) and pd.isna(original_uncurated_cohort)) and curated_cohort != original_uncurated_cohort:
                    df.at[index, FileMetadata.COHORT_YG_AY] = curated_cohort
                    logger.info(
                        f"The Cohort of '{row[FileMetadata.FILE_NAME]} - {row[FileMetadata.SHEET_NAME]}' "
                        f"has been updated from {original_uncurated_cohort} to {curated_cohort}"
                    )
                    updates_made = True

    # Process files with standardised filenames, if applicable.
    # The standardised filenames must follow one of these formats:
    # - For NCCIS data: {data category}-{y11 cohort or recorded month and year}.xlsx 
    #   (e.g., post16Dest_2024-Mar.xlsx)
    # - General format: {data category}-{y11 cohort}_{year group}-{corresponding cohort of this year group}.xlsx 
    #   (e.g., activitySurvey_2018-2019_in_Y11-2018-2019.xlsx)
    # The order of the data category, y11 cohort, year group, and year cohort can be different and set in file_naming_format.

    if has_processed:
        raise NotImplementedError(
            "Processing files with standardised filenames is not yet fully implemented by `extract_file_metadata()`. "
            "Please use `extract_file_metadata_from_folder()` instead."
        )
        df[FileMetadata.CATEGORY] = df[FileMetadata.FILE_NAME].apply(lambda x: x.split('_')[0])
        
        df[FileMetadata.COHORT_Y11_AY] = df[FileMetadata.FILE_NAME].apply(
            lambda x: x.split("_")[1].replace("-", " - ").rsplit(".", 1)[0]
        )
     
        df[FileMetadata.COHORT_Y11_AY] = df[FileMetadata.COHORT_Y11_AY].apply(
            lambda x: ' '.join(x.split(' - ')) if len(x.split(' - ')) > 1 and x.split(' - ')[1].isalpha() else x
        )
        
        df[FileMetadata.YEAR_GRP] = df[FileMetadata.FILE_NAME].apply(
            lambda x: None if "_in_" not in x else (x.split("_in_")[1].split("-")[0][1:] if len(x.split("_in_")) > 1 else None)
        )
        
        df[FileMetadata.COHORT_YG_AY] = df[FileMetadata.FILE_NAME].apply(
            lambda x: None if "_in_" not in x else x.split("_in_")[1].rsplit(".", 1)[0]
        )
        
        df[FileMetadata.COHORT_YG_AY] = df[FileMetadata.COHORT_YG_AY].apply(
            lambda x: f"{x.split('-')[0]} {x.split('-')[1]} - {x.split('-')[2]}" if x is not None else x
        )
        
        df[FileMetadata.NEEDS_REVIEW] = None
    
    # Check for duplicates based on [CATEGORY, COHORT_Y11_AY, YEAR_GRP]
    dup_mask = df.duplicated(subset=[FileMetadata.CATEGORY, FileMetadata.COHORT_Y11_AY, FileMetadata.YEAR_GRP], keep=False)
    dup_mask = dup_mask & (df[FileMetadata.CATEGORY] != FileMetadata.EXCLUDE)
    if dup_mask.any():
        # If duplicates found, print the duplicate rows using tabulate and raise an error
        logger.warning(
            f"Duplicate entries found in file metadata based on [{FileMetadata.CATEGORY}, {FileMetadata.COHORT_Y11_AY}, {FileMetadata.YEAR_GRP}]. "
            f"Please ensure only one table is used per cohort per {'/'.join(valid_cat)}. "
            f"Duplicate rows have been highlighted in red in the Excel file. Please resolve them before proceeding (e.g., change corresponding 'Category' as {FileMetadata.EXCLUDE})."
        )
        duplicate_rows = df[dup_mask]
        
        # Set NEEDS_REVIEW to "YES" for duplicate rows
        df.loc[dup_mask, FileMetadata.NEEDS_REVIEW] = "YES"
        
        print(tabulate(duplicate_rows, headers='keys', tablefmt='fancy_outline', showindex=False))
    
    # Print all rows that need review
    needs_review_rows = df[df[FileMetadata.NEEDS_REVIEW].str.upper() == "YES"]
    if not needs_review_rows.empty:
        logger.warning("Rows that need review:")
        print(tabulate(needs_review_rows, headers='keys', tablefmt='fancy_outline', showindex=False))
        
    # Get the index of rows that need review
    needs_review_index = needs_review_rows.index
    
    # Highlight rows
    def highlight_rows(row):
        if row.name in needs_review_index:
            return ['background-color: #FFB6C1'] * len(row)
        return [''] * len(row)

    df_styled = df.style.apply(highlight_rows, axis=1)
    
    # Save updated metadata
    if new_entries_added or updates_made:
        if len(needs_review_index) > 0:
            df_styled.to_excel(output_path, index=False, engine="openpyxl")
        else:
            df.to_excel(output_path, index=False)
        logger.info(f"{output_path} has been updated successfully.")
    else:
        logger.info(f"No new entries or updates to add to {output_path}.")

    cat_summary = data_upload_report(df)

    # Generate and display upload report
    if display_upload_status:
        missing_report = cat_summary['missing_report']
        styled_print(f"Upload Data Summary - Missing Data ({len(missing_report)} records):")
        print(tabulate(missing_report, headers='keys', tablefmt='fancy_outline', showindex=False))

        styled_print("Upload Data Summary:")
        print(tabulate(cat_summary['summary'], headers='keys', tablefmt='fancy_outline', showindex=False))

    # Save full report
    full_report = cat_summary['full_report']
    full_report.to_excel(output_path.parent / 'report_data_upload_status.xlsx', index=False)

    styled_print("File metadata extraction process completed.", colour="magenta")

    return df, cat_summary


def add_col_info(
    col_metadata: Dict[str, pd.DataFrame],
    yaml_data: Dict[str, Dict[str, Union[str, List[Dict[str, str]]]]],
    input_path: Union[str, Path],
    output_path: Union[str, Path],
) -> Dict[str, pd.DataFrame]:
    """
    Append descriptions and values information to each sheet in col_metadata from multiple metadata sources.
    Adds a suffix from the nickname if there are multiple sources, but removes duplicates if identical.

    Parameters
    ----------
    col_metadata : dict
        Dictionary where keys are sheet names and values are DataFrames of the column metadata for each category.
                
    yaml_data : dict, 
        Dictionary containing metadata file information and mappings for additional columns.
        Example structure:
            attendance:  # Data category
              is_meta: False
              metadata:
                sources:
                  - filename: School data metadata (census, attendance, exclusions).xlsx
                    sheetname: Attendance
                    colname: Field
                    description: Info
                    value: Values
                
    input_path : Union[str, Path]
        Directory containing the metadata files.
        
    output_path : Union[str, Path]
        Path where the Excel file should be saved.

    Returns
    -------
    dict
        Updated col_metadata with appended description and values information for each sheet.
    """
    
    # Keep a copy of the old col_metadata for comparison later
    old_col_metadata = {s: df.copy() for s, df in col_metadata.items()}

    for data_category, config in sorted(yaml_data.items()):
        # Skip categories without metadata sources or those marked as metadata
        if 'metadata' not in config or config.get('is_meta', False):
            continue

        # Retrieve the sheet DataFrame for the data category
        if data_category not in col_metadata:
            col_metadata[data_category] = pd.DataFrame(columns=[ColumnMetadata.STD_NAME])
        
        sheet_df = col_metadata[data_category]

        # Determine the join key: use "Source Column" if it exists, otherwise "Column Name"
        # "Source Column" is the raw column name (without standarisation)
        join_column = (
            ColumnMetadata.SRC_NAME 
            if ColumnMetadata.SRC_NAME in sheet_df.columns 
            else ColumnMetadata.STD_NAME
        )
        
        if join_column not in sheet_df.columns:
            print(f"Neither {ColumnMetadata.SRC_NAME} nor {ColumnMetadata.STD_NAME} found in {data_category} sheet.")
            continue
        
        description_columns = []
        value_columns = []
        
        # Process each source under metadata
        for source in config['metadata']['sources']:
        
            # Define the full path to the file and check if it exists
            file_path = input_path / source['filename']
            if not file_path.exists():
                print(f"File not found: {file_path}")
                continue

            # Load the specific sheet if provided, else default to the first sheet
            sheetname = source.get('sheetname', 0)
            skiprows = source.get('skiprows', None)
            try:
                df_metadata = pd.read_excel(file_path, sheet_name=sheetname, skiprows=skiprows)
            except Exception as e:
                print(f"Error reading file {file_path}: {e}")
                continue

            # Check for required columns in the metadata file
            colname = source['colname']
            description_col = source['description']
            value_col = source['value']
                
            required_columns = [colname, description_col, value_col]
            missing_columns = [col for col in required_columns if col not in df_metadata.columns]

            if missing_columns:
                print(f"Missing required columns in {file_path} sheet {sheetname}: {', '.join(missing_columns)}")
                continue
            
            # Set nickname suffix for columns
            nickname = source.get('nickname', '')
            description_col_name = (
                f"{ColumnMetadata.DESCRIPTION}_{nickname}" if nickname else ColumnMetadata.DESCRIPTION
            )
            value_col_name = f"{ColumnMetadata.VALUES}_{nickname}" if nickname else ColumnMetadata.VALUES

            # Track description and value for duplicate check
            description_columns.append(description_col_name)
            value_columns.append(value_col_name)

            # Ensure the columns with suffix exist
            if description_col_name not in sheet_df.columns:
                sheet_df[description_col_name] = None
            if value_col_name not in sheet_df.columns:
                sheet_df[value_col_name] = None

            for _, row in df_metadata.iterrows():
                key_value = row[colname]
                description = row[description_col]
                value = row[value_col]
                    
                # Filter sheet_df for columns that match 'key_value'
                mask = sheet_df[join_column] == key_value

                # Update existing entries with new descriptions and values
                if mask.any():
                    sheet_df.loc[mask, description_col_name] = description
                    sheet_df.loc[mask, value_col_name] = value
            
        # Check for identical 'Description' columns
        if len(description_columns) > 1:
            filled_desc_df = sheet_df[description_columns].fillna('')
            
            if filled_desc_df.nunique(axis=1).eq(1).all():
                # All descriptions are identical; merge into a single column
                sheet_df[ColumnMetadata.DESCRIPTION] = sheet_df[description_columns[0]]
                sheet_df.drop(columns=description_columns, inplace=True)
            else:
                # Add a column to show specific differences in descriptions
                sheet_df['Description_Differences'] = filled_desc_df.apply(get_differences, axis=1)

        # Check for identical 'Values' columns
        if len(value_columns) > 1:
            filled_values_df = sheet_df[value_columns].fillna('')
            
            if filled_values_df.nunique(axis=1).eq(1).all():
                # All values are identical; merge into a single column
                sheet_df['Values'] = sheet_df[value_columns[0]]
                sheet_df.drop(columns=value_columns, inplace=True)
            else:
                # Add a column to show specific differences in values
                sheet_df['Values_Differences'] = filled_values_df.apply(get_differences, axis=1)
        
            # Update col_metadata with the modified sheet
            col_metadata[data_category] = sheet_df
    
    # Check if any changes have been made to col_metadata
    has_changes = compare_and_backup(
        new_data=col_metadata,
        old_data=old_col_metadata,
        target_path=output_path,
        backup_prefix="precurate",
        save_new_data=False,
    ) 
    
    if not has_changes:
        return col_metadata

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for sheet_name, df in col_metadata.items():
            cols_to_remove = [ColumnMetadata.DETECTED_DTYPES]
            df = df[[col for col in df.columns if col not in cols_to_remove]]

            # Check if the columns exist, and move them to the end
            cols_to_move = [ColumnMetadata.STD_FILE, ColumnMetadata.SRC_FILE]
            for col in cols_to_move:
                if col in df.columns and col not in cols_to_remove:
                    df = df[[c for c in df.columns if c != col] + [col]]
            
            # Writing each DataFrame to a separate sheet
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    styled_print(f"Descriptions and Values have been appended", colour="magenta")
        
    return col_metadata

# not necessay especially when using df = df.convert_dtypes(dtype_backend="pyarrow")
def curate_col_dtype(
    col_metadata: Dict[str, pd.DataFrame],
    standardise_rules: list[dict],
    output_path: Union[str, Path],
) -> Dict[str, pd.DataFrame]:
    """
    Curate the data types for specified columns in col_metadata and update with curated data types.
    Highlight cells in the "Curated Data Type" column if they differ from "Uncurated Data Type".

    Parameters
    ----------
    col_metadata : dict
        Dictionary where keys are sheet names and values are DataFrames of the column metadata for each category.
    
    standardise_rules : list[dict]
        List of rename actions applied in order.  
        Each item is one of:
        * **{"replace": <str>, "to": <str>}** - simple find-&-substitute.  
        * **{"add_underscore_before_caps": true}** - insert "_" before capital letters.  
        * **{"to_lower": true}** - convert to lower-case.

        Omit for no renaming (defaults to []).

    output_path : Union[str, Path]
        Path where the updated Excel file should be saved. Defaults to self.col_metadata_path.

    Returns
    -------
    dict
        Updated col_metadata with curated data types and additional metadata columns.
    """
   
    # Keep a copy of old col_metadata for comparison
    old_col_metadata = {s: df.copy() for s, df in col_metadata.items()}        
    
    highlighted_output_path = output_path.with_name(f"{output_path.stem}{'_highlighted'}{output_path.suffix}")

    # Create a workbook object
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)  # Remove the default empty sheet

    # Red fill for highlighting
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    
    from .naming import standardise_colnames
    
    # Columns that need to be curated
    (
        string_cols,
        unordered_category_cols,
        ordered_category_cols,  # Not fully implemented
        datetime_cols,
        numeric_cols,
        float_cols,
    ) = (
        standardise_colnames(cols, standardise_rules)
        for cols in (
            ColumnSchema.STRING_COLS,
            ColumnSchema.UNORDERED_CATEGORY_COLS,
            ColumnSchema.ORDERED_CATEGORY_COLS,
            ColumnSchema.DATETIME_COLS,
            ColumnSchema.NUMERIC_COLS,
            ColumnSchema.FLOAT_COLS,
        )
    )
    
    column_type_mapping = {
        "string": set(string_cols),
        "category": set(unordered_category_cols) | set(ordered_category_cols),
        "datetime": set(datetime_cols),
        "Float64": set(float_cols),
        "numeric": set(numeric_cols),
    }
    
    for sheet_name, sheet_df in sorted(col_metadata.items()):
        # Ensure required columns exist or create them
        for required_col in [
            ColumnMetadata.DATA_TYPE,
            ColumnMetadata.CURATED_DTYPE,
            ColumnMetadata.RAW_DTYPE,
            ColumnMetadata.UNIQUE_COUNT,
            ColumnMetadata.UNIQUE_VALUES,
        ]:
            if required_col not in sheet_df.columns:
                sheet_df[required_col] = None

        # Merge all columns with prefix "Values" into a single "Values" column
        value_cols = [
            col for col in sheet_df.columns 
            if col.startswith(f"{ColumnMetadata.VALUES}_") and not col.endswith("Differences")
        ]
        if value_cols:
            sheet_df[ColumnMetadata.VALUES] = sheet_df[value_cols].apply(
                lambda row: '\n'.join(dict.fromkeys(
                    filter(None, (re.sub(r'[^\w\s]+$', '', str(item)).strip() if pd.notna(item) else None for item in row))
                )), axis=1
            )
            # sheet_df.drop(columns=value_cols, inplace=True)
            
        # Merge all columns with prefix "Description" into a single "Description" column
        desp_cols = [
            col for col in sheet_df.columns 
            if col.startswith(f"{ColumnMetadata.DESCRIPTION}_") and not col.endswith("Differences")
        ]
        if desp_cols:
            sheet_df[ColumnMetadata.DESCRIPTION] = sheet_df[desp_cols].apply(
                lambda row: '\n'.join(dict.fromkeys(
                    filter(None, (re.sub(
                        r'[^\w\s]+$', '', str(item)).strip() if pd.notna(item) else None for item in row))
                )), axis=1
            )
            # sheet_df.drop(columns=desp_cols, inplace=True)
        
        if ColumnMetadata.RAW_DTYPE in sheet_df.columns:
            # The data type has already been curated before, need to check if any updates made
            old_uncurated_dtype = sheet_df[ColumnMetadata.RAW_DTYPE].copy()
            old_curated_dtype = sheet_df[ColumnMetadata.CURATED_DTYPE].copy()
        
        # Copy original data types to "Uncurated Data Type"
        sheet_df[ColumnMetadata.RAW_DTYPE] = sheet_df[ColumnMetadata.DATA_TYPE]
        sheet_df[ColumnMetadata.CURATED_DTYPE] = sheet_df[ColumnMetadata.DATA_TYPE]
        
        # Curate data types for each column
        for idx, row in sheet_df.iterrows():
            col_name = row[ColumnMetadata.STD_NAME]
            unique_values = row[ColumnMetadata.UNIQUE_VALUES]
            if not pd.isna(unique_values):
                parsed_unique_values = {v.strip().lower() for v in unique_values.split(",")}
                # num_unique_values = row[ColumnMetadata.UNIQUE_COUNT]
                values_mapping = row[ColumnMetadata.VALUES]
                
                assigned_dtypes = set()
                
                # Check for partial matches in any category
                for dtype, keyword_set in column_type_mapping.items():
                    if col_name in keyword_set or remove_prefix(col_name) in keyword_set:
                        assigned_dtypes.add(dtype)
                
                # Special handling for boolean detection
                if len(parsed_unique_values) > 1 and parsed_unique_values.issubset(ColumnSchema.BOOLEAN_VALUES):
                    assigned_dtypes.add("boolean")
                elif isinstance(values_mapping, str) and (
                    all(x in values_mapping.lower() for x in ["0 = no", "1 = yes"]) or
                    all(x in values_mapping.lower() for x in ["0 = false", "1 = true"])
                ):
                    assigned_dtypes.add("boolean")
                
                if len(assigned_dtypes) > 1:
                    raise ValueError(
                        f"Conflict detected: Column '{col_name}' was assigned multiple curated dtypes: {assigned_dtypes}"
                    )
                
                if assigned_dtypes:
                    sheet_df.at[idx, ColumnMetadata.CURATED_DTYPE] = assigned_dtypes.pop()
          
        sheet_df[ColumnMetadata.CURATED_DTYPE].replace('timestamp', 'datetime', inplace=True)
        sheet_df[ColumnMetadata.CURATED_DTYPE].replace('int64', 'Int64', inplace=True)
        sheet_df[ColumnMetadata.CURATED_DTYPE].replace('double', 'Float64', inplace=True)
        sheet_df[ColumnMetadata.CURATED_DTYPE].replace('bool', 'boolean', inplace=True)
        # sheet_df[ColumnMetadata.CURATED_DTYPE].replace('float64', 'Float64', inplace=True)

        # Ensure the final "Data Type" matches the curated data type
        sheet_df[ColumnMetadata.DATA_TYPE] = sheet_df[ColumnMetadata.CURATED_DTYPE]
        
        if (sheet_df[ColumnMetadata.DATA_TYPE] == old_curated_dtype).all():
            sheet_df[ColumnMetadata.RAW_DTYPE] = old_uncurated_dtype

        # Reorder the columns
        desired_order = [
            ColumnMetadata.STD_NAME,
            ColumnMetadata.SRC_NAME,
            ColumnMetadata.DATA_TYPE,
            ColumnMetadata.CURATED_DTYPE,
            ColumnMetadata.RAW_DTYPE,
            ColumnMetadata.DESCRIPTION,
            ColumnMetadata.VALUES,
        ]
        
        ordered_columns = [col for col in desired_order if col in sheet_df.columns]
        remaining_columns = [col for col in sheet_df.columns if col not in ordered_columns]
        sheet_df = sheet_df[ordered_columns + remaining_columns]
        
        # Update the sheet in col_metadata
        col_metadata[sheet_name] = sheet_df

        # Write the sheet to the workbook
        ws = wb.create_sheet(title=sheet_name)
        for row_idx, row in enumerate(dataframe_to_rows(sheet_df, index=False, header=True), start=1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Highlight if the data type has been curated/changed
                # col_idx and row_idx are 1-based indexing (Excel's indexing) 
                # col_idx - 1: converts to Python's 0-based indexing
                # row_idx - 2: converts to Python's 0-based indexing and skips header row
                if row_idx > 1 and sheet_df.columns[col_idx - 1] == ColumnMetadata.CURATED_DTYPE:
                    uncurated_value = sheet_df.at[row_idx - 2, ColumnMetadata.RAW_DTYPE]
                    curated_value = value
                    if curated_value != uncurated_value:
                        cell.fill = red_fill

    # Check if changes occurred
    has_changes = compare_and_backup(
        new_data=col_metadata,
        old_data=old_col_metadata,
        target_path=output_path,
        backup_prefix="precurate",
        save_new_data=False,
    ) 
    
    if not has_changes:
        return col_metadata

    # Save the workbook (highlighted version, for further review)
    wb.save(highlighted_output_path)
    
    # save cleaned version
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for sheet_name, df in col_metadata.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    styled_print(
        f"Saved curated data:\n"
        f"- Highlighted version (curated entries marked in red): {highlighted_output_path}\n"
        f"- Cleaned version: {output_path}",
        colour='magenta'
    )
    
    return col_metadata
