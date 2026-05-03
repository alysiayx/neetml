import pickle
import logging
import re
import warnings
import yaml
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import itertools
import textwrap
import seaborn as sns
import matplotlib.pyplot as plt
from matplotlib.colors import Colormap
import difflib
from typing import Union, Optional
from pathlib import Path
from termcolor import colored
from difflib import SequenceMatcher
from tabulate import tabulate
from rich.console import Console
from rich.text import Text
from rich.table import Table
from rich.panel import Panel
from typing import List, Set
from prettytable import PrettyTable, ALL
from itertools import combinations
from collections import defaultdict

def set_default_path(
    path: Union[str, Path, None],
    default_path: Union[str, Path],
) -> Path:
    """
    Set the path and create directories if necessary. If the provided path is None, use the default path.
    Automatically detect whether the path is a directory or a file.

    Parameters
    ----------
    path : Union[str, Path, None]
        The path to check. If None, the default path is used.
    
    default_path : Union[str, Path]
        The default path to use if the provided path is None.

    Returns
    -------
    Path
        The final path that was set.
    """
    if path is None:
        path = Path(default_path)
    else:
        path = Path(path)

    # Detect if the path is a directory by checking if it has a file extension
    if path.suffix == '':
        # No file extension means it's a directory, so create the directory
        path.mkdir(parents=True, exist_ok=True)

    return path

def styled_print(message: str, colour: str = 'cyan', bold: bool = True, console: Console = None):
    """
    Print a message with a specified color and optional bold formatting.

    Parameters
    ----------
    message : str
        The message to print.
    color : str, optional
        The color to use for the message, by default 'cyan'.
    bold : bool, optional
        Flag to indicate if the message should be printed in bold, by default True.
    console : Console, optional
        A rich Console object for advanced rendering, by default None.
    """
    if console is None:
        # Use termcolor to style the text in the terminal
        attributes = ['bold'] if bold else []
        print(colored(message, color=colour, attrs=attributes))
    else:
        # Use rich for styling and rendering
        text = Text(message, style=colour)
        if bold:
            text.stylize("bold")
        console.print(text)

def load_dataframe(
    file_path: Union[str, Path],
    sheet_name: Union[str, int, None] = None,
    **kwargs
) -> pd.DataFrame:
    """
    Load a DataFrame from a given CSV, XLSX, or Parquet file path.
    
    Parameters
    ----------
    file_path : Union[str, Path]
        Path to the CSV, XLSX, or Parquet file.

    sheet_name : Union[str, int, None], optional
        Name or index of the sheet to load for Excel files. If None, the first sheet will be loaded.

    **kwargs : dict
        Additional keyword arguments passed directly to `pd.read_csv()`, `pd.read_excel()`, or `pd.read_parquet()`.
    
    Returns
    -------
    pd.DataFrame
        The loaded DataFrame.

    Raises
    ------
    ValueError
        If the file is not in a supported format (CSV, XLSX, or Parquet).
    """
    file_path = Path(file_path)

    if file_path.suffix == '.xlsx':
        return pd.read_excel(file_path, sheet_name=sheet_name or 0, **kwargs)
    elif file_path.suffix == '.csv':
        return pd.read_csv(file_path, **kwargs)
    elif file_path.suffix == '.parquet':
        return pd.read_parquet(file_path, **kwargs)
    else:
        raise ValueError("Unsupported file format. Please provide a CSV, XLSX, or Parquet file.")

def resolve_dataframe(
    df: pd.DataFrame | None = None,
    path: str | Path | None = None,
    default_path: str | Path | None = None,
    name: str = "data",
    copy: bool = True,
    warn: bool = True,
    logger: logging.Logger | None = None
) -> pd.DataFrame:
    """
    Resolve a dataframe from:
    1) a provided DataFrame,
    2) a provided file path,
    3) a default file path.
    """
    if df is not None:
        if path is not None and warn:
            warnings.warn(
                f"Both '{name}' and '{name}_path' were provided. "
                f"Using '{name}' and ignoring '{name}_path'.",
                stacklevel=2,
            )
        logger.debug(f"Using provided DataFrame for {name}.") if logger else None
        styled_print(f"Using provided DataFrame for {name}.")
        return df.copy(deep=True) if copy else df

    resolved_path = path or default_path
    if resolved_path is None:
        raise ValueError(
            f"No {name} provided. Please provide either '{name}' or '{name}_path'."
        )

    resolved_path = Path(resolved_path)
    if not resolved_path.exists():
        raise FileNotFoundError(
            f"File not found: {resolved_path}"
        )
    else:
        logger.debug(f"Loading {name} from: {resolved_path}") if logger else None
        styled_print(f"Loading {name} from: {resolved_path}")

    return load_dataframe(resolved_path)

def parse_yaml(file_path: Union[str, Path]) -> dict:
    # Load YAML configuration file and return its content as a dictionary
    if not isinstance(file_path, Path):
        file_path = Path(file_path)

    with file_path.open('r') as file:
        data = yaml.safe_load(file)

    return data

def get_files_in_folder(
    folder_path: Union[str, Path], 
    recursive: bool = True, 
    extensions: Set[str] = {".xlsx", ".csv"}
) -> List[Path]:
    """
    List all files in the specified folder with user-defined extensions.
    
    Parameters
    ----------
    folder_path : Union[str, Path]
        The folder where files are located.

    recursive : bool, optional
        If True (default), searches recursively using `rglob`. 
        If False, only lists files in the top-level directory using `glob`.

    extensions : Set[str], optional
        A set of file extensions to filter (default: {".xlsx", ".csv"}).
        Example: {".parquet", ".json"} to list Parquet and JSON files.

    Returns
    -------
    List[Path]
        A sorted list of file paths.
    """
    folder_path = Path(folder_path)
    
    if isinstance(extensions, str):
        extensions = {extensions} 

    # Choose between glob (non-recursive) and rglob (recursive)
    search_method = folder_path.rglob if recursive else folder_path.glob

    files = sorted([
        file for file in search_method("*") 
        if file.suffix.lower() in extensions and not file.name.startswith((".", "~"))
    ])
    
    return files

def list_sheets_in_excel(file_path: Union[str, Path]) -> list:
    # List all sheets in the specified Excel file and sort them alphabetically
    file_path = Path(file_path)
    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheets = sorted(wb.sheetnames)
    wb.close()
    return sheets

def display_columns_table(columns_set: Set[str], title: str = None, max_columns: int = 10) -> None:
    """
    Display a table of columns with a specified title. Columns are arranged in a table format 
    with a specified maximum number of columns per row.

    Parameters
    ----------
    columns_set : Set[str]
        A set of column names to display in the table.
        
    title : str, optional
        The title of the table to display.
        
    max_columns : int, optional
        Maximum number of columns per row (default is 10).
    """
    columns_list = sorted(columns_set)
    num_columns = len(columns_list)
    
    # Calculate the number of rows needed based on max_columns
    num_rows = (num_columns + max_columns - 1) // max_columns
    
    # Pad the list to make it divisible by max_columns
    padded_columns_list = columns_list + [''] * (num_rows * max_columns - num_columns)
    
    # Reshape the list into a 2D array with max_columns columns
    columns_array = np.array(padded_columns_list).reshape(num_rows, max_columns)
    
    # Create DataFrame
    columns_df = pd.DataFrame(columns_array)
    
    # Display the table with title
    if title:
        print(f"\n{title}:")
    
    print(tabulate(columns_df, tablefmt='simple', showindex=False))

def stringify_unique_values(values, max_values=10):
    """
    Converts a set of unique values into a string representation.

    Parameters
    ----------
    values : set
        Set of unique values to be converted to string.

    max_values : int, optional
        Maximum number of unique values to include in the string. Additional values are represented by '...'.

    Returns
    -------
    str
        String representation of the unique values.
    """
    values = {'Missing' if pd.isna(value) else value for value in values}
    values_list = list(values)
    try:
        values_list.sort()
    except TypeError:
        # Cannot sort values of different types
        pass
    values_list = values_list[:max_values]
    values_str = ', '.join(map(str, values_list))
    if len(values) > max_values:
        values_str += ', ...'
    return values_str

def similar_score(a: str, b: str) -> float:
    """
    Calculate the similarity ratio between two strings.

    Parameters
    ----------
    a : str
        The first string for comparison.
        
    b : str
        The second string for comparison.

    Returns
    -------
    float
        The similarity ratio between the two strings.
    """
    return SequenceMatcher(None, a, b).ratio()

def get_base_name(col_name: str) -> str:
    """
    Extract the base name from a column name, removing numeric suffixes if present.

    Parameters
    ----------
    col_name : str
        The column name to process.

    Returns
    -------
    str
        The base name of the column without numeric suffixes.
    """
    # get the base name of a column name
    # Split on the last occurrence of '.', if it's followed by digits
    if '.' in col_name and col_name.rsplit('.', 1)[1].isdigit():
        return col_name.rsplit('.', 1)[0]
    else:
        return col_name

def add_underscore_before_caps(s: str) -> str:
    """
    Add underscores before uppercase letters in a string, except when they are consecutive uppercase letters.

    Parameters
    ----------
    s : str
        The string to modify.

    Returns
    -------
    str
        The modified string with underscores added before uppercase letters.
    """
    # If the string already contains an underscore, return it as is
    if '_' in s:
        return s
    
    # If the entire string is uppercase, return it as is
    if s.isupper():
        return s
    
    # This regex identifies where to add underscores:
    # - (?<!^) ensures no underscore is added at the start
    # - (?<![A-Z]) prevents underscores between consecutive uppercase letters
    # - (?=[A-Z]) adds an underscore before uppercase letters starting a new word
    return re.sub(r'(?<!^)(?<![A-Z])(?=[A-Z])', '_', s)

def has_mixed_case_no_whitespace(s: str) -> bool:
    """
    Check if a string has both uppercase and lowercase letters and contains no whitespace.

    Parameters
    ----------
    s : str
        The string to check.

    Returns
    -------
    bool
        True if the string has mixed case with no whitespace, otherwise False.
    """
    # Check if the string has more than one uppercase and more than one lowercase letter with no whitespace
    return bool(re.search(r'(?=(.*[A-Z].*){2,})(?=(.*[a-z].*){2,})^\S+$', s))

def get_differences(row: pd.Series) -> str:
    """
    Identify differences between unique, non-NaN values in a row of columns.

    Parameters
    ----------
    row : pd.Series
        The row of values to compare for differences.

    Returns
    -------
    str
        The differences between values, formatted with + or - to indicate changes, or an empty string if no differences.
    """
    unique_values = row.dropna().unique()  # Get unique, non-NaN values
    if len(unique_values) < 2:
        return ''  # No differences if all values are the same
    
    # Compare the first and last unique values for simplicity
    # TODO: can be improved
    diff = difflib.ndiff(unique_values[0].splitlines(), unique_values[-1].splitlines())
    return '\n'.join([line for line in diff if line.startswith('+ ') or line.startswith('- ')])

def visible_files_count(folder: Path) -> int:
    return sum(1 for file in folder.glob('*') if not file.name.startswith('.'))

def check_folder_file_count_equal(folder1: Path, folder2: Path) -> bool:
    """
    Check if the number of non-hidden files in two folders is equal.

    Args:
        folder1 (Path): The path to the first folder.
        folder2 (Path): The path to the second folder.

    Returns:
        bool: True if the number of non-hidden files in both folders is equal, False otherwise.
    """

    return visible_files_count(folder1) == visible_files_count(folder2)

def print_table(
    df: pd.DataFrame,
    console: Console = None,
    rich_table: bool = True,
    num_rows: Union[int, str] = 5,
    num_cols: Union[int, str] = 10,
    num_records: Union[int, str] = 2,
    id_column: str = "stud_id",
    group_records: bool = True,
    show_index: bool = False,
    show_diffs_only: bool = False,
    fill_na: bool = True,
    title: str = None,
    wrap_in_panel: bool = False,  # Option to wrap the table inside a Panel
) -> None:
    """
    Display a DataFrame in a tabular format with grouping and formatting options.

    If some rows or columns are hidden due to display limits, append a summary row/column indicating the count of hidden elements.

    Parameters
    ----------
    df : pd.DataFrame
        The DataFrame to be displayed.
    console : Console, optional
        The Rich console to use for rendering (default is None, will be created if needed).
    rich_table : bool, optional
        Whether to use Rich Table (default is True).
    num_rows : int or "all", optional
        Number of rows to display (default is 5, "all" to show all).
    num_cols : int or "all", optional
        Number of columns to display (default is 10, "all" to show all).
    num_records : int or "all", optional
        Number of unique records to show in grouping mode (default is 2, "all" to show all).
    id_column : str, optional
        The column used for grouping records (default is "stud_id").
    group_records : bool, optional
        Whether to group records by the ID column (default is True).
    show_index : bool, optional
        Whether to include index as a column (default is False).
    show_diffs_only : bool, optional
        Whether to show only columns with unique values (default is False).
    title : str, optional
        Title to display above the table (default is None).
    wrap_in_panel : bool, optional
        Whether to wrap the table inside a Rich Panel (default is False).
    fill_na : bool, optional
        Whether to replace NaN values with empty strings (default is True).
    """
    
    data = df.copy()
    
    if not console and rich_table:
        console = Console()
    
    if fill_na:
        # # Handle categorical columns safely
        # for col in data.select_dtypes(include="category"):
        #     if "" not in data[col].cat.categories:
        #         data[col] = data[col].cat.add_categories("")
        #     data[col] = data[col].fillna("")

        # # Fill NaNs for all other non-numeric columns
        # data = data.apply(lambda col: col.fillna("") if col.dtype.kind not in "biufc" else col)
        
        # # Fill NaNs for all columns (including numeric) with an empty string
        # for col in data.columns:
        #     if pd.api.types.is_numeric_dtype(data[col]):
        #         data[col] = data[col].astype(object).fillna("")
        #     else:
        #         # For other columns, simply fill NaN with an empty string
        #         data[col] = data[col].fillna("")
        data = data.astype(object).fillna("")
    
    # Include the index as a column if requested
    if show_index:
        index_name = data.index.name or "Index"
        data = data.reset_index(names=index_name)

    # Handle 'all' for num_rows, num_cols, and num_records
    total_rows, total_cols = len(data), len(data.columns)
    num_rows = total_rows if num_rows == "all" else min(num_rows, total_rows)
    num_cols = total_cols if num_cols == "all" else min(num_cols, total_cols)
    num_records = (
        data[id_column].nunique()
        if num_records == "all" and id_column in data.columns
        else num_records
    )

    def add_hidden_summary(data, full_rows, full_cols):
        """Add summary for hidden rows/columns."""
        hidden_row_count = full_rows - len(data)
        hidden_col_count = full_cols - len(data.columns)

        # Add a row indicating hidden rows
        if hidden_row_count > 0:
            hidden_row = pd.DataFrame(
                [["... Hidden rows: " + str(hidden_row_count)] + [""] * (len(data.columns) - 1)],
                columns=data.columns
            )
            data = pd.concat([data, hidden_row], ignore_index=True)

        # Add a column indicating hidden columns
        if hidden_col_count > 0:
            summary_column = pd.Series(
                ["... Hidden columns: " + str(hidden_col_count)] * len(data),
                name=""
            )
            data = pd.concat([data, summary_column], axis=1)

        return data

    def render_table(data, table_title=None):
        """Render table using Rich.Table or PrettyTable based on user preference."""
        if not rich_table or data.shape[1] > 11:  # Use PrettyTable if rich_table=False or too many columns
            table = PrettyTable()
            table.field_names = data.columns.tolist()
            table.max_width = 15
            table.hrules = ALL  # Add horizontal rules for better readability
            for _, row in data.iterrows():
                table.add_row(row.tolist())
            print(f"{table_title or ''}\n{table}\n")
        
        else:  # Use Rich.Table
            table = Table(title=table_title if not wrap_in_panel else None, highlight=True, expand=True)
            for col in data.columns:
                table.add_column(col, no_wrap=False, overflow="fold")
            for _, row in data.iterrows():
                row_str = [str(item) if pd.notna(item) else "" for item in row]
                table.add_row(*row_str)

            # Wrap table in a panel if the user opts in
            if wrap_in_panel:
                table = Panel(table, title=table_title, border_style="cyan")

            console.print(table)

    def filter_differences(data):
        """Remove duplicated columns."""
        return data.loc[:, data.nunique(axis=0, dropna=False) > 1]

    # General table summary
    if not id_column or id_column not in data.columns or not group_records:
        if show_diffs_only:
            data = filter_differences(data)
        df_limited = data.iloc[:num_rows, :num_cols]
        df_limited = add_hidden_summary(df_limited, total_rows, total_cols)
        render_table(df_limited, table_title=title)
        return

    # Grouped table rendering
    df_grouped = data.groupby(id_column)
    for group_key, group in itertools.islice(df_grouped, num_records):
        group_limited = group.iloc[:num_rows, :num_cols].dropna(axis=1, how="all")
        if show_diffs_only:
            group_limited = filter_differences(group_limited)
        render_table(group_limited, table_title=f"{id_column.capitalize()}: {group_key}")

def assert_no_object_columns(df: pd.DataFrame):
    """
    Checks if the provided DataFrame contains any object-type columns.
    Raises a TypeError with details if such columns exist.

    Parameters:
        df (pd.DataFrame): The DataFrame to check.

    Raises:
        TypeError: If object-type columns are found.
    """
    object_cols = df.select_dtypes(include='object').columns.tolist()
    if object_cols:
        raise TypeError(f"Object-type columns detected: {', '.join(object_cols)}")

# def save_model(model, filename):
#     """Save the model to disk."""
#     with open(filename, 'wb') as file:
#         pickle.dump(model, file)

# def load_model(filename):
#     """Load a model from disk."""
#     with open(filename, 'rb') as file:
#         return pickle.load(file)

# def load_config():
#     """Load configuration from YAML file."""
#     with open("config/config.yaml", "r") as file:
#         config = yaml.safe_load(file)
#     return config

## slow version
# def find_comp_pairs(df, cols=None, thres=0.6, skip_prefixes=('ks2', 'ks4')):
#     """
#     Find column pairs that are:
#     - Name-similar
#     - Strictly complementary (per row, at most one non-null)
#     - Same dtype, or both categorical (category/object)
    
#     Parameters:
#         df: pandas DataFrame
#         cols: optional list of columns to consider; if None, uses all except 'stud_id' and underscore-prefixed
#         thres: similarity threshold for column name comparison
#         skip_prefixes: tuple of prefixes to skip warning when category values differ

#     Returns:
#         complementary_pairs: list of (colA, colB) tuples
#     """
#     if cols is None:
#         cols = df.columns

#     comp_pairs = []

#     for colA, colB in combinations(cols, 2):
#         # Check name similarity
#         if similar_score(colA, colB) < thres:
#             continue

#         # Check strictly complementary
#         pair_df = df[[colA, colB]].dropna(how='all')
#         max_non_na_per_row = pair_df.nunique(dropna=True, axis=1).max()
#         if max_non_na_per_row > 1:
#             continue

#         dtypeA, dtypeB = df[colA].dtype, df[colB].dtype

#         # If dtype not same, check if both are categorical
#         if dtypeA != dtypeB:
#             dtypes = df[[colA, colB]].dtypes
#             if all(dt.name == 'category' for dt in dtypes):
#                 if all(col.lower().startswith(skip_prefixes) for col in [colA, colB]):
#                     continue
#                 else:
#                     print(f"⚠️ '{colA}' and '{colB}' are both categorical, but have different category values.")
#             else:
#                 continue

#         print(f"✅ '{colA}' and '{colB}' are strictly complementary, name-similar, and have same dtype ({dtypeA})")
#         comp_pairs.append((colA, colB))

#     return comp_pairs

## speedy version
def find_comp_pairs(df, cols=None, thres=0.6, skip_prefixes=('ks2', 'ks4')):
    """
    Faster version of find_comp_pairs.

    Returns
    -------
    comp_pairs : list[tuple[str, str]]
        Same logic as original function.
    minor_to_major : dict[str, str]
        Map from minority column -> majority column,
        based on non-null count among each complementary pair.
    """
    if cols is None:
        cols = sorted(list(df.columns))

    comp_pairs = []
    minor_to_major = {}

    dtypes = df[cols].dtypes.to_dict()
    non_null_counts = df[cols].notna().sum().to_dict()
    lower_cols = {c: c.lower() for c in cols}

    count = 0
    for colA, colB in combinations(cols, 2):
        # name similarity
        # only compare columns if are similar enough to avoid unnecessary computations
        if similar_score(colA, colB) < thres:
            continue

        dtypeA, dtypeB = dtypes[colA], dtypes[colB]

        # dtype check
        if dtypeA != dtypeB:
            if dtypeA.name == 'category' and dtypeB.name == 'category':
                if all(lower_cols[c].startswith(skip_prefixes) for c in (colA, colB)):
                    continue
                # else:
                #     print(f"⚠️ '{colA}' and '{colB}' are both categorical, but have different category values.")
            else:
                continue

        # strict complement check
        sA = df[colA]
        sB = df[colB]
        overlap = sA.notna() & sB.notna()

        if overlap.any():
            a = sA[overlap].astype("object")
            b = sB[overlap].astype("object")
            if not a.eq(b).all():
                continue
        
        count += 1
        # print(f"{count}: '{colA}' and '{colB}' are strictly complementary, name-similar, and have same dtype ({dtypeA})")
        comp_pairs.append((colA, colB))

        # minority -> majority
        cntA = non_null_counts[colA]
        cntB = non_null_counts[colB]

        if cntA < cntB:
            minor_to_major[colA] = colB
        elif cntB < cntA:
            minor_to_major[colB] = colA
        else:
            # if counts are equal, use alphabetical order to decide
            minor_to_major[min(colA, colB)] = max(colA, colB)

    return comp_pairs, minor_to_major

def preview_col_map(df, mapping):
    rows = []

    for from_col, to_col in mapping.items():
        rows.append({
            "from_col": from_col,
            "to_col": to_col,
            "from_non_null": int(df[from_col].notna().sum()) if from_col in df.columns else None,
            "to_non_null": int(df[to_col].notna().sum()) if to_col in df.columns else None,
            "dtype_from": str(df[from_col].dtype) if from_col in df.columns else None,
            "dtype_to": str(df[to_col].dtype) if to_col in df.columns else None,
        })

    return pd.DataFrame(rows)

def apply_col_map(df, mapping, drop_from_col=True, copy=True, verbose=True):
    if copy:
        df = df.copy()

    if verbose:
        preview_df = preview_col_map(df, mapping)
        print("\n Preview of column mapping:")
        print(tabulate(preview_df, headers='keys', tablefmt='psql', showindex=False))    
    
    for from_col, to_col in mapping.items():
        if from_col == to_col:
            continue
        if from_col not in df.columns:
            continue
        if to_col not in df.columns:
            if from_col in df.columns:
                df.rename(columns={from_col: to_col}, inplace=True) # if to_col doesn't exist, just rename from_col to to_col
            continue

        if pd.api.types.is_categorical_dtype(df[to_col]) and pd.api.types.is_categorical_dtype(df[from_col]):
            new_cats = df[to_col].cat.categories.union(df[from_col].cat.categories)
            df[to_col] = df[to_col].cat.set_categories(new_cats)
            df[to_col] = df[to_col].combine_first(df[from_col])
        else:
            df[to_col] = df[to_col].combine_first(df[from_col]).astype(str(df[to_col].dtype)) # fill to_col with from_col values where to_col is null

        if drop_from_col:
            df.drop(columns=from_col, inplace=True)

    return df

def plot_col_coverage(
    df: pd.DataFrame,
    prefix: str = "attendance",
    max_year_group: int = 11,
    cmap: Union[str, Colormap] = "Blues",
    figsize: Optional[tuple] = None,
    output_file: str = "cols_presence.jpg",
    dpi: int = 800,
    verbose: bool = False,
    save_excel: bool = False,
):
    """
    Generate a heatmap showing the non-missing column coverage of columns
    across cohort-year groups, to reflect where variables have actual values
    rather than structural presence only.

    Parameters
    ----------
    df : pd.DataFrame
        Input DataFrame containing student records and cohort/year group columns.
    prefix : str, optional
        Prefix to filter columns of interest (default: "attendance").
    max_year_group : int, optional
        Only include year groups less than or equal to this value (default: 11).
    cmap : str or matplotlib.colors.Colormap, optional
        Colormap for the heatmap (default: "Blues").
    figsize : tuple, optional
        Size of the matplotlib figure (default: None).
    output_file : str, optional
        Path to save the generated heatmap image (default: "cols_presence.jpg").
    dpi : int, optional
        Resolution (dots per inch) for the saved image (default: 800).
    verbose : bool, optional
        If True, display the plot interactively; otherwise, save and close (default: False).
    save_excel: bool, option
        If True, export heatmap in Excel (default: False)

    Returns
    -------
    None
        The function saves the heatmap to the specified file and optionally displays it.

    Notes
    -----
    - Requires columns: 'stud_id', '_y11_cohort', '_year_group' in the DataFrame.
    - Each cell in the heatmap indicates whether a column is present (non-empty) for a given (cohort, year group).
    - Useful for data auditing and feature engineering.
    """
    df = df.copy()

    selected_cols = [col for col in df.columns if col.startswith(prefix)]
    df_selected = df[selected_cols + ['stud_id', '_cohort_y11_ay', '_year_group']]
    
    print(f"Number of students: {df['stud_id'].nunique()}")
    print(f"Columns with prefix '{prefix}': {len(selected_cols)}")
    if len(selected_cols) == 0:
        print(f"No columns with prefix '{prefix}' found in the data")
        return
    
    print(f"Shape of filtered df: {df_selected.shape}")

    col_to_groups = defaultdict(set)
    all_groups = []

    for cohort_year, group in df_selected.groupby(['_cohort_y11_ay', '_year_group']):
        if int(cohort_year[1]) > max_year_group:
            continue

        group = group.dropna(axis=1, how='all')
        cols = set(group.columns) - {'_cohort_y11_ay', '_year_group', 'stud_id'}
        for col in cols:
            col_to_groups[col].add(cohort_year)

        all_groups.append(cohort_year)

    total_groups = len(all_groups)

    summary = pd.DataFrame([
        {
            "column": col,
            "n_groups": len(groups),
            "proportion": len(groups) / total_groups,
            "groups": sorted(groups)
        }
        for col, groups in col_to_groups.items()
    ]).sort_values("proportion")

    all_groups = sorted(all_groups, key=lambda x: (x[0], int(x[1])))
    all_groups_str = [f"{c}_Y{y}" for (c, y) in all_groups]

    presence = pd.DataFrame(0, index=summary["column"], columns=all_groups_str)

    for col, groups in col_to_groups.items():
        for (c, y) in groups:
            colname = f"{c}_Y{y}"
            presence.loc[col, colname] = 1

    n_rows = presence.shape[0]
    n_cols = presence.shape[1]
    height = max(10, min(n_rows * 0.2, 20))
    width  = max(10, min(n_cols * 0.4, 30))
    
    if figsize is None:
        plt.figure(figsize=(width, height))
    else:
        plt.figure(figsize=figsize)
    
    sns.heatmap(
        presence,
        cmap=cmap,
        cbar=False,
        linewidths=0.5,
        linecolor="white"
    )
    plt.xlabel("(cohort_year)")
    plt.ylabel("Column")
    plt.title(f"Non-missing column coverage across cohort-year groups (Y<={max_year_group})")
    
    if "ks4" in prefix:
        ax = plt.gca()
        note = (
            "*Note: KS4 data is collected in year 11 but only becomes available from year 12 onward."
        )
        plt.gcf().text(
            ax.get_position().x0,      
            0.01,         
            note,
            fontsize=10,
            style="italic",
            ha="left"
        )

    plt.savefig(output_file, dpi=dpi, bbox_inches="tight")
    if verbose:
        plt.show()
    else:
        plt.close()
    print(f"Heatmap saved to {output_file}")

    if save_excel:
        # export presence to Excel
        excel_file = str(output_file).replace(".jpg", ".xlsx")
        presence.to_excel(excel_file)

        # open workbook
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        # define a green fill for cells with 1
        green_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # apply conditional format
        for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if cell.value == 1:
                    cell.fill = green_fill
                    cell.value = ""
                else:
                    cell.value = ""

        wb.save(excel_file)
        print(f"Excel presence matrix saved to {excel_file} with highlights.")
        
def plot_group_heatmap(
    df: pd.DataFrame,
    index_col: str,
    columns_col: str,
    values_col: str,
    x_label: Optional[str] = None,
    y_label: Optional[str] = None,
    aggfunc: str = "nunique",
    title: str = "Heatmap",
    cmap: str = "Blues",
    figsize: tuple = (12, 8),
    output_file: Optional[str] = None,
    annot: bool = True,
    fmt: str = ".0f"
):
    """
    Plots a heatmap showing aggregation of `values_col` across (index_col, columns_col) pairs.
    
    Parameters
    ----------
    df : pd.DataFrame
        Input data
    index_col : str
        Column to place on the heatmap y-axis
    columns_col : str
        Column to place on the heatmap x-axis
    values_col : str
        Column to aggregate
    x_label : str, optional
        Label for the x-axis
    y_label : str, optional
        Label for the y-axis
    aggfunc : str, optional
        Aggregation function ('nunique', 'count', etc.), by default "nunique"
    title : str, optional
        Title of the heatmap
    cmap : str, optional
        Color map, by default "Blues"
    figsize : tuple, optional
        Figure size, by default (12,8)
    output_file : str, optional
        If provided, saves the figure
    annot : bool, optional
        Whether to annotate heatmap cells
    fmt : str, optional
        Format of annotations
    """
    if aggfunc == "nunique":
        grouped = (
            df.groupby([index_col, columns_col])[values_col]
            .nunique()
            .reset_index()
        )
    elif aggfunc == "count":
        grouped = (
            df.groupby([index_col, columns_col])[values_col]
            .count()
            .reset_index()
        )
    elif aggfunc == "sum":
        grouped = (
            df.groupby([index_col, columns_col])[values_col]
            .sum()
            .reset_index()
        )
    else:
        raise ValueError(f"Unsupported aggfunc: {aggfunc}")
    
    heatmap_data = grouped.pivot(
        index=index_col,
        columns=columns_col,
        values=values_col
    )

    # Sort the columns (x-axis) numerically if possible, otherwise lexicographically
    try:
        sorted_columns = sorted(heatmap_data.columns, key=lambda x: int(x))
    except (ValueError, TypeError):
        sorted_columns = sorted(heatmap_data.columns)
    heatmap_data = heatmap_data[sorted_columns]

    plt.figure(figsize=figsize)
    sns.heatmap(
        heatmap_data,
        annot=annot,
        fmt=fmt,
        cmap=cmap
    )
    plt.title(title)
    plt.xlabel(x_label if x_label else columns_col)
    plt.ylabel(y_label if y_label else index_col)
    plt.tight_layout()
    if output_file:
        plt.savefig(output_file)
        print(f"Saved heatmap to {output_file}")
    plt.show()

